#-*- coding:utf-8 -*-
from rdkit import Chem
import openpyxl
import os, sys

from rdkit.Chem import rdCoordGen #cyclic to 각지게
#rdCoordGen.AddCoords(mol)

def get_positions(ws):
    positions = {}

    for c in range(1, ws.max_column + 1):
        positions[ws.cell(row = 1, column = c).value] = c
        
    return positions


def write_info(src, dest, smis, ws_title=None, codes=None, names=None, mws=None, srcs=None): #template path, dest path filename
    wb = openpyxl.load_workbook(f"{src}")
    ws = wb.worksheets[0]
    if ws_title: ws.title = ws_title

    positions = get_positions(ws)

    for row_num in range(1, len(smis)+1):
        if row_num == 1: 
            continue
        
        ws.cell(row = row_num, column = positions['No.']).value = row_num - 1
        ws.cell(row = row_num, column = positions['SMILES']).value = smis[row_num-2]

        if names: 
            ws.cell(row = row_num, column = positions['Name']).value = names[row_num-2]
        if mws:
            ws.cell(row = row_num, column = positions['MW']).value = mws[row_num-2]
        if codes:
            ws.cell(row = row_num, column = positions['Code']).value = codes[row_num-2]
        if srcs:
            ws.cell(row = row_num, column = positions['Source']).value = srcs[row_num-2]

    save_xl(wb, dest)
    print(ws.cell(row = row_num, column = positions['SMILES']).value)

def save_xl(wb, dest):
    while True:
        try:
            wb.save(f"{dest}")
            break
        except:
            continue

def draw_img(src, dest):
    os.makedirs('./img', exist_ok=True)
    wb = openpyxl.load_workbook(f"{src}")
    ws = wb.worksheets[0]
    loglines = []

    positions = get_positions(ws)
    ws.column_dimensions[chr(65 + positions['Structure']-1)].width = 32
    print(positions)
    
    for i, row in enumerate(ws.iter_rows()):
        if i == 0: continue
        print(f'{i/ws.max_row:.3f}% {i}')
        mol = Chem.MolFromSmiles(row[positions['SMILES']-1].value)
        if not mol:
            loglines.append(f'Invalid SMILES: {row[positions["SMILES"]-1].value}')
            print(loglines)
            continue

        img_name = f'./img/{i}.png'
        #rdCoordGen.AddCoords(mol)
        #Chem.Draw.MolToFile(mol, img_name, size=(500, 300))
        Chem.Draw.MolToFile(mol, img_name, size=(250, 150))
        print(Chem.MolToSmiles(mol))

        img = openpyxl.drawing.image.Image(img_name)
        for cell in row:
            if cell.coordinate[0] != chr(65 + positions['Structure']-1): continue
            img.anchor = cell.coordinate
            ws.add_image(img)

        ws.row_dimensions[i+1].height = 120

    save_xl(wb, dest)

    with open('./draw_img_log.txt', 'w') as f:
        f.write('\n'.join(loglines))
    print('done')